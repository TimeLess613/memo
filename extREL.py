#!/usr/bin/env python
# coding=utf-8
import os
import shutil
import datetime
import time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

from gotoweb import service_now

print("############################################")
print("################  START  ###################")
service_now()


####  datetime for fileName  ####

today = datetime.datetime.now().strftime("%Y%m%d")
todayMMYY = datetime.datetime.now().strftime("%m%d")
recent = [(datetime.datetime.now() + datetime.timedelta(days=-i)).strftime("%Y%m%d") for i in range(1, 31)]
print("--------",today)
print("--------",todayMMYY)

####  fileFormat   ####
colWidth = [12, 27, 50, 10, 10, 10, 20, 20, 30, 6]
setFill = PatternFill("solid", fgColor="6495ED")
setBorder = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

####  get log  ####
def getlog():
    getUser = os.getcwd()
    getTime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    myPath = "//Uninas05.uninas.mew.co.jp/un1345i/share/E10/02.個人管理フォルダ/CYatpan/0.張/00.home/python/extREL/"
    logFile = myPath + "log_getlog.txt"
    with open(logFile, 'a') as f:
       f.write(getUser + ':  ' + getTime + '\n')
getlog()

####  Set FileEnv  ####
fileSysPath = "//uninas05.uninas.mew.co.jp/UN1345I/share/E10/01.事業部管理フォルダ/LSVA関係作業共有/ランチミーティングリスト/"
print("fileSysPath: " + fileSysPath)

downloadFile = "C:/Users/4079157/Downloads/change_task.xlsx"    # for localAdmin
# downloadFile = fileSysPath + "change_task.xlsx"    # for otherUser
print("Edit downloadFile: " + downloadFile)
wbOrig = load_workbook(downloadFile)
wsOrig = wbOrig.active
wsOrig.title = todayMMYY
wsOrig['A1'].value = "番号" + todayMMYY
for i in range(1, wsOrig.max_column+1):
    wsOrig.column_dimensions[get_column_letter(i)].width = colWidth[i-1]
wbOrig.save(downloadFile)
wbOrig.close()

changedFileName = fileSysPath + "NW作業一覧(" + today + "抜出).xlsx"
shutil.move(downloadFile, changedFileName)
print("--------","Moved downloadFile to fileSys")    # for localAdmin
# print("Changed downloadFileName ")    # for otherUser

fileNew = changedFileName
def find_recentfile():
    for i in range(len(recent)):
        fileOld = fileSysPath + "NW作業一覧(" + recent[i] + "抜出).xlsx"
        if os.path.exists(fileOld):
            print("fileOld: " + fileOld)
            return fileOld
fileOld = find_recentfile()
print("fileNew: " + fileNew)

####  Get REL Diff to List ####
print("--------","Get REL Diff to List")
wbNew = load_workbook(fileNew)
wsNew = wbNew.active
listNew = [item.value for item in list(wsNew.columns)[0]]
wbNew.close()

wbOld = load_workbook(fileOld)
wsOld = wbOld.active
listOld = [item.value for item in list(wsOld.columns)[0]]
wbOld.close()

listDiff = set(listNew) - set(listOld)
print(listDiff)

####  Find NewREL in fileNew ####
dumSheet = []
for r in range(1, wsNew.max_row+1):
    dumRow = []
    for c in range(1, wsNew.max_column+1):
        if wsNew.cell(r, 1).value in listDiff:
            dumRow.append(wsNew.cell(r,c).value)
    dumSheet.append(dumRow)
sheet = [x for x in dumSheet if x != []]

####  Set Update Excel  ####
print("--------","Set Update Excel")
wb = Workbook()
ws = wb.active
ws.title = todayMMYY

row = len(sheet)
col = len(sheet[0])
for i in range(1, row+1):
    for j in range(1, col+1):
        ws.cell(i, j).value = sheet[i-1][j-1]
        ws.cell(i, j).fill = setFill
        ws.cell(i, j).border = setBorder 
for i in range(1, col):
    ws.column_dimensions[get_column_letter(i)].width = colWidth[i-1]

wb.close()
wb.save(fileSysPath + "NW作業一覧(" + today + "更新).xlsx")
print(fileSysPath + "NW作業一覧(" + today + "更新).xlsx")
print("#################  DONE  ###################")
print("############################################")